from datetime import date

from django.utils.translation import gettext
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.writer.excel import save_virtual_workbook

from drf_renderer_xlsx.renderers import XLSXRenderer, get_attribute, get_style_from_dict


class CustomXLSXRenderer(XLSXRenderer):
    """
    Custom Renderer for Excel spreadsheet open data format (xlsx) for override the format_number of date cells, to be
    written on the excel file as dates cells
    """

    media_type = "application/xlsx"
    format = "xlsx"
    row_color = "row_color"

    def render(self, data, accepted_media_type=None, renderer_context=None):  # noqa C901
        """
        Render `data` into XLSX workbook, returning a workbook.
        """
        if not self._check_validatation_data(data):
            return self._json_format_response(data)

        if data is None:
            return bytes()

        wb = Workbook()
        ws = wb.active

        results = data["results"] if "results" in data else data

        # Take header and column_header params from view
        header, header_title = self.get_header(ws, renderer_context)
        header_style = get_style_from_dict(header.get("style"), "header_style")

        column_header = get_attribute(renderer_context["view"], "column_header", {})
        column_header_style = get_style_from_dict(
            column_header.get("style"), "column_header_style"
        )

        column_count = 0
        row_count = 1
        if header:
            row_count += 1
        # Make column headers
        column_titles = column_header.get("titles", [])

        # If we have results, pull the columns names from the keys of the first row
        if len(results):
            column_names_first_row = self._flatten(results[0])
            for column_name in column_names_first_row.keys():
                if column_name == self.row_color:
                    continue
                column_count += 1
                column_name_display = self.select_column_name(column_name, column_count, column_titles)
                ws.cell(row=row_count, column=column_count, value=column_name_display).style = column_header_style
            ws.row_dimensions[row_count].height = column_header.get("height", 45)

        # Set the header row
        self.set_header_row(ws, header, column_count, header_title, header_style)

        # Set column width
        self.set_column_width(ws, column_header, column_count)

        # Make body
        body = get_attribute(renderer_context["view"], "body", {})
        body_style = get_style_from_dict(body.get("style"), "body_style")
        self.fill_body(ws, results, body, body_style, row_count)

        return save_virtual_workbook(wb)

    def get_header(self, ws, renderer_context):
        # Take header and column_header params from view
        header = get_attribute(renderer_context["view"], "header", {})
        ws.title = header.get("tab_title", "Report")
        header_title = header.get("header_title", "Report")
        img_addr = header.get("img")
        if img_addr:
            img = Image(img_addr)
            ws.add_image(img, "A1")
        return header, header_title

    def select_column_name(self, column_name, column_count, column_titles):
        return gettext(column_name) if column_count > len(column_titles) else column_titles[column_count - 1]

    def set_columns_names(self, ws, column_names_first_row, column_count, column_titles, column_header_style,
                          row_count):
        for column_name in column_names_first_row.keys():
            if column_name == self.row_color:
                continue
            column_count += 1
            column_name_display = self.select_column_name(column_name, column_count, column_titles)
            ws.cell(row=row_count, column=column_count, value=column_name_display).style = column_header_style

    @staticmethod
    def set_header_row(ws, header, column_count, header_title, header_style):
        if header:
            last_col_letter = "G"
            if column_count:
                last_col_letter = get_column_letter(column_count)
            ws.merge_cells("A1:{}1".format(last_col_letter))

            cell = ws.cell(row=1, column=1, value=header_title)
            cell.style = header_style
            ws.row_dimensions[1].height = header.get("height", 45)

    @staticmethod
    def set_column_width(ws, column_header, column_count):
        column_width = column_header.get("column_width", 20)
        if isinstance(column_width, list):
            for i, width in enumerate(column_width):
                col_letter = get_column_letter(i + 1)
                ws.column_dimensions[col_letter].width = width
        else:
            for ws_column in range(1, column_count + 1):
                col_letter = get_column_letter(ws_column)
                ws.column_dimensions[col_letter].width = column_width

    def fill_body(self, ws, results, body, body_style, row_count):
        for row in results:
            column_count = 0
            row_count += 1
            flatten_row = self._flatten(row)
            for column_name, value in flatten_row.items():
                if column_name == self.row_color:
                    continue
                column_count += 1
                try:
                    cell = ws.cell(row=row_count, column=column_count, value=value)
                except IllegalCharacterError:
                    value = value.replace("\r", " ").replace("\n", " ").replace("\v", "")
                    cell = ws.cell(row=row_count, column=column_count, value=value)
                cell.style = body_style
                # Override numberformat of date cells
                if isinstance(value, date):
                    cell.number_format = FORMAT_DATE_DDMMYY
            ws.row_dimensions[row_count].height = body.get("height", 40)

            self.fill_row_background(ws, row, column_count, row_count)

    def fill_row_background(self, ws, row, column_count, row_count):
        if self.row_color in row:
            last_letter = get_column_letter(column_count)
            cell_range = ws["A{}".format(row_count): "{}{}".format(last_letter, row_count)]
            fill = PatternFill(fill_type="solid", start_color=row[self.row_color])
            for r in cell_range:
                for c in r:
                    c.fill = fill
